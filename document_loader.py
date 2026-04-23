import logging
import os
from pathlib import Path
from typing import Any, Dict, Iterator, List

import pdfplumber
import docx
import openpyxl

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {'.pdf', '.docx', '.txt', '.md', '.xlsx'}


def normalize_text(text: str) -> str:
    return text.replace('\r', '\n').replace('\t', ' ').strip()


def iter_text_segments(root_path: str) -> Iterator[Dict[str, Any]]:
    root = Path(root_path)
    for path in sorted(root.rglob('*')):
        if not path.is_file():
            continue
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue

        try:
            if path.suffix.lower() == '.pdf':
                yield from parse_pdf(path)
            elif path.suffix.lower() == '.docx':
                yield from parse_docx(path)
            elif path.suffix.lower() in {'.txt', '.md'}:
                yield parse_text_file(path)
            elif path.suffix.lower() == '.xlsx':
                yield from parse_xlsx(path)
        except Exception as exc:
            logger.warning(f'Failed to parse {path}: {exc}', exc_info=True)


def parse_pdf(path: Path) -> Iterator[Dict[str, Any]]:
    with pdfplumber.open(path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            text = normalize_text(page.extract_text() or '')
            if not text:
                continue
            yield {
                'source': path.name,
                'content': text,
                'metadata': {
                    'file_path': str(path),
                    'file_name': path.name,
                    'page': page_number,
                    'modified_time': path.stat().st_mtime,
                },
            }


def parse_docx(path: Path) -> Iterator[Dict[str, Any]]:
    document = docx.Document(path)
    for paragraph_index, paragraph in enumerate(document.paragraphs, start=1):
        text = normalize_text(paragraph.text)
        if not text:
            continue
        yield {
            'source': path.name,
            'content': text,
            'metadata': {
                'file_path': str(path),
                'file_name': path.name,
                'paragraph': paragraph_index,
                'modified_time': path.stat().st_mtime,
            },
        }


def parse_text_file(path: Path) -> Dict[str, Any]:
    with path.open('r', encoding='utf-8', errors='ignore') as f:
        text = normalize_text(f.read())
    return {
        'source': path.name,
        'content': text,
        'metadata': {
            'file_path': str(path),
            'file_name': path.name,
            'modified_time': path.stat().st_mtime,
        },
    }


def parse_xlsx(path: Path) -> Iterator[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                rows.append(' '.join(values))
        text = normalize_text('\n'.join(rows))
        if not text:
            continue
        yield {
            'source': f'{path.name}#{sheet_name}',
            'content': text,
            'metadata': {
                'file_path': str(path),
                'file_name': path.name,
                'sheet_name': sheet_name,
                'modified_time': path.stat().st_mtime,
            },
        }
